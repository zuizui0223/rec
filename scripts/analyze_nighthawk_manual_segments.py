#!/usr/bin/env python3
"""Audit released manual versus Nighthawk segment counts without event-level overclaiming."""
from __future__ import annotations

import argparse
import csv
import json
import math
from pathlib import Path
from typing import Any, Iterable


class NighthawkAuditError(ValueError):
    pass


MANUAL_ALIASES = {
    "manualcalls",
    "manualcall",
    "manualcount",
    "manualcounts",
    "manual_call_count",
    "manual_calls",
}
NIGHTHAWK_ALIASES = {
    "nighthawkcalls",
    "nighthawkcall",
    "nighthawkcount",
    "nighthawkcounts",
    "nighthawk_call_count",
    "nighthawk_calls",
    "modelcalls",
    "modelcount",
}
FILENAME_ALIASES = {"filename", "file", "recording", "recordingfile"}
DURATION_ALIASES = {"duration", "durationseconds", "duration_seconds"}


def _canon(name: object) -> str:
    return "".join(ch.lower() for ch in str(name).strip() if ch.isalnum() or ch == "_")


def _find_column(headers: list[str], aliases: set[str], label: str) -> str:
    by_canon = {_canon(h): h for h in headers}
    hits = [by_canon[a] for a in aliases if a in by_canon]
    if len(hits) != 1:
        raise NighthawkAuditError(
            f"expected exactly one {label} column matching {sorted(aliases)}, found {hits}"
        )
    return hits[0]


def _optional_column(headers: list[str], aliases: set[str]) -> str | None:
    by_canon = {_canon(h): h for h in headers}
    hits = [by_canon[a] for a in aliases if a in by_canon]
    return hits[0] if len(hits) == 1 else None


def _number(value: object, label: str) -> float:
    if value is None or str(value).strip() == "":
        raise NighthawkAuditError(f"{label} is missing")
    try:
        out = float(value)
    except (TypeError, ValueError) as exc:
        raise NighthawkAuditError(f"{label} must be numeric, got {value!r}") from exc
    if not math.isfinite(out) or out < 0:
        raise NighthawkAuditError(f"{label} must be finite and >=0, got {value!r}")
    return out


def _load_csv(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    with path.open(newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        if reader.fieldnames is None:
            raise NighthawkAuditError("CSV has no header")
        return list(reader.fieldnames), [dict(r) for r in reader]


def _load_xlsx(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise NighthawkAuditError("XLSX input requires openpyxl") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        try:
            first = next(rows)
        except StopIteration as exc:
            raise NighthawkAuditError("XLSX is empty") from exc
        headers = [str(v).strip() if v is not None else "" for v in first]
        if not any(headers):
            raise NighthawkAuditError("XLSX has blank header")
        out: list[dict[str, object]] = []
        for values in rows:
            if all(v is None or str(v).strip() == "" for v in values):
                continue
            out.append({h: (values[i] if i < len(values) else None) for i, h in enumerate(headers)})
        return headers, out
    finally:
        wb.close()


def load_table(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _load_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _load_xlsx(path)
    raise NighthawkAuditError(f"unsupported table type {suffix!r}; expected CSV/XLSX")


def _pearson(x: list[float], y: list[float]) -> float | None:
    if len(x) != len(y) or len(x) < 2:
        return None
    mx = sum(x) / len(x)
    my = sum(y) / len(y)
    dx = [v - mx for v in x]
    dy = [v - my for v in y]
    den = math.sqrt(sum(v * v for v in dx) * sum(v * v for v in dy))
    return sum(a * b for a, b in zip(dx, dy)) / den if den > 0 else None


def _ranks(values: list[float]) -> list[float]:
    order = sorted(range(len(values)), key=values.__getitem__)
    ranks = [0.0] * len(values)
    i = 0
    while i < len(order):
        j = i + 1
        while j < len(order) and values[order[j]] == values[order[i]]:
            j += 1
        avg = (i + 1 + j) / 2.0
        for k in range(i, j):
            ranks[order[k]] = avg
        i = j
    return ranks


def _spearman(x: list[float], y: list[float]) -> float | None:
    return _pearson(_ranks(x), _ranks(y))


def _ols(x: list[float], y: list[float]) -> tuple[float | None, float | None]:
    if len(x) != len(y) or not x:
        return None, None
    mx = sum(x) / len(x)
    my = sum(y) / len(y)
    ssx = sum((v - mx) ** 2 for v in x)
    if ssx <= 0:
        return None, my
    slope = sum((a - mx) * (b - my) for a, b in zip(x, y)) / ssx
    return slope, my - slope * mx


def _quantile(values: list[float], p: float) -> float:
    vals = sorted(values)
    if not vals:
        raise NighthawkAuditError("cannot compute quantile of empty values")
    if len(vals) == 1:
        return vals[0]
    pos = p * (len(vals) - 1)
    lo = int(math.floor(pos))
    hi = int(math.ceil(pos))
    if lo == hi:
        return vals[lo]
    f = pos - lo
    return vals[lo] * (1 - f) + vals[hi] * f


def _error_summary(pairs: list[tuple[float, float]]) -> dict[str, float | int | None]:
    if not pairs:
        return {
            "n": 0,
            "manual_total": 0.0,
            "nighthawk_total": 0.0,
            "recorded_to_manual_total_ratio": None,
            "mean_signed_error": None,
            "mae": None,
            "rmse": None,
        }
    manual = [a for a, _ in pairs]
    model = [b for _, b in pairs]
    errors = [b - a for a, b in pairs]
    mt = sum(manual)
    nt = sum(model)
    return {
        "n": len(pairs),
        "manual_total": mt,
        "nighthawk_total": nt,
        "recorded_to_manual_total_ratio": nt / mt if mt > 0 else None,
        "mean_signed_error": sum(errors) / len(errors),
        "mae": sum(abs(e) for e in errors) / len(errors),
        "rmse": math.sqrt(sum(e * e for e in errors) / len(errors)),
    }


def analyze(headers: list[str], rows: list[dict[str, object]]) -> dict[str, Any]:
    if not rows:
        raise NighthawkAuditError("validation table contains no rows")
    manual_col = _find_column(headers, MANUAL_ALIASES, "manual-count")
    model_col = _find_column(headers, NIGHTHAWK_ALIASES, "Nighthawk-count")
    filename_col = _optional_column(headers, FILENAME_ALIASES)
    duration_col = _optional_column(headers, DURATION_ALIASES)

    pairs: list[tuple[float, float]] = []
    filenames: list[str | None] = []
    durations: list[float | None] = []
    for i, row in enumerate(rows, start=2):
        manual = _number(row.get(manual_col), f"row {i} {manual_col}")
        model = _number(row.get(model_col), f"row {i} {model_col}")
        pairs.append((manual, model))
        filenames.append(None if filename_col is None else str(row.get(filename_col, "")).strip() or None)
        durations.append(None if duration_col is None else _number(row.get(duration_col), f"row {i} {duration_col}"))

    manual = [a for a, _ in pairs]
    model = [b for _, b in pairs]
    slope, intercept = _ols(manual, model)
    positive = [(a, b) for a, b in pairs if a > 0]
    zero = [(a, b) for a, b in pairs if a == 0]

    quartiles: dict[str, Any] = {}
    if positive:
        positive_manual = [a for a, _ in positive]
        q1, q2, q3 = (_quantile(positive_manual, p) for p in (0.25, 0.5, 0.75))
        bins: dict[str, list[tuple[float, float]]] = {"Q1": [], "Q2": [], "Q3": [], "Q4": []}
        for pair in positive:
            a = pair[0]
            key = "Q1" if a <= q1 else "Q2" if a <= q2 else "Q3" if a <= q3 else "Q4"
            bins[key].append(pair)
        quartiles = {
            "manual_positive_cutpoints": {"q25": q1, "q50": q2, "q75": q3},
            "groups": {key: _error_summary(vals) for key, vals in bins.items()},
        }

    overall = _error_summary(pairs)
    overall.update(
        {
            "pearson_r": _pearson(manual, model),
            "spearman_rho": _spearman(manual, model),
            "ols_slope_nighthawk_on_manual": slope,
            "ols_intercept_nighthawk_on_manual": intercept,
        }
    )

    zero_summary = _error_summary(zero)
    zero_summary["segments_with_positive_nighthawk_count"] = sum(1 for _, b in zero if b > 0)
    zero_summary["fraction_with_positive_nighthawk_count"] = (
        zero_summary["segments_with_positive_nighthawk_count"] / len(zero) if zero else None
    )

    return {
        "schema": "rec-nighthawk-white-sands-count-audit-v1",
        "column_mapping": {
            "manual_count": manual_col,
            "nighthawk_count": model_col,
            "filename": filename_col,
            "duration": duration_col,
        },
        "overall": overall,
        "manual_zero_segments": zero_summary,
        "manual_positive_segments": _error_summary(positive),
        "positive_manual_count_quartiles": quartiles,
        "row_provenance": {
            "row_count": len(rows),
            "filename_available": filename_col is not None,
            "duration_available": duration_col is not None,
            "unique_nonblank_filenames": len({f for f in filenames if f}) if filename_col else None,
            "duration_seconds_min": min((d for d in durations if d is not None), default=None),
            "duration_seconds_max": max((d for d in durations if d is not None), default=None),
        },
        "identification_boundary": (
            "This table compares aggregate segment counts. Count differences cannot identify event-level missed-call "
            "or false-entry prevalence because omissions and extra entries may cancel within a segment. Do not use "
            "this output to estimate q_shadow or a_R. The admissible REC claim is segment-level ecological-count "
            "preservation/distortion (H3 boundary)."
        ),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("table", type=Path)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    try:
        result = analyze(*load_table(args.table))
    except NighthawkAuditError as exc:
        raise SystemExit(f"Nighthawk manual audit failed: {exc}") from exc
    text = json.dumps(result, indent=2, sort_keys=True) + "\n"
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(text, encoding="utf-8")
    print(text, end="")


if __name__ == "__main__":
    main()
